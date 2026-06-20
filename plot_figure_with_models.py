"""
Plot neuromechanical model coordinates from an editable Excel workbook.

Input workbook:
    /mnt/data/neuromechanical_models_full_table_tickspace.xlsx

Coordinate convention:
    - workbook coordinates run from 0 to 1 between the first and last tick.
    - the first and last ticks are offset inside the physical axis limits.
    - each model receives exactly one projection to each of:
      controller, actuator, and N DOF.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import to_rgb
from matplotlib.patches import FancyArrowPatch, Polygon


MODEL_XLSX = Path("models_data.xlsx")

OUT_PNG = Path("figure_with_models.png")
OUT_PDF = Path("figure_with_models.pdf")


# =============================================================================
# Style
# =============================================================================

plt.rcParams.update(
    {
        "font.family"     : "sans-serif",
        "font.sans-serif" : ["Arial", "Helvetica", "DejaVu Sans"],
        "pdf.fonttype"    : 42,
        "ps.fonttype"     : 42,
    }
)

ACTUATOR_COLOR          = "#3b4fd1"
NDOF_COLOR              = "#0087d7"
CONTROLLER_COLOR        = "#7b4ab2"
ENVIRONMENT_COLOR       = "#2f8f46"
BOX_COLOR               = "#6f6f6f"

ROBOT_COLOR             = "#d00000"
MODEL_COLOR             = "#000000"

AXIS_LINEWIDTH          = 2.2
BOX_LINEWIDTH           = 0.72
PROJECTION_LINEWIDTH    = 0.68
VERTICAL_LINEWIDTH      = 1.1
READOUT_AXIS_LINEWIDTH  = 1.35

SPHERE_ALPHA            = 0.65
VERTICAL_LINE_ALPHA     = 0.65
PROJECTION_ALPHA        = 0.78


# =============================================================================
# Geometry
# =============================================================================

ORIGIN          = np.array([1.25, 2.58])
EMBODIMENT_VEC  = np.array([7.75, 0.00])
CONTROLLER_VEC  = np.array([2.35, -2.66])
ENVIRONMENT_VEC = np.array([0.00, 3.40])

NDOF_OFFSET     = 0.08
NDOF_START      = ORIGIN + NDOF_OFFSET * CONTROLLER_VEC

AXIS_TICK_START = 0.12
AXIS_TICK_END   = 0.90

FIGSIZE         = (12.2, 7.6)
DPI             = 160
X_LIMITS        = (0.0, 14.10)
Y_LIMITS        = (-0.80, 7.10)


def clip_01(value: float) -> float:
    return min(max(float(value), 0.0), 1.0)


def axis_pos(coord: float) -> float:
    """Map semantic coordinate 0..1 to an offset axis position."""
    return AXIS_TICK_START + clip_01(coord) * (AXIS_TICK_END - AXIS_TICK_START)


def project_point(
    controller  : float = 0.0,
    embodiment  : float = 0.0,
    environment : float = 0.0,
) -> np.ndarray:
    return (
        ORIGIN
        + controller  * CONTROLLER_VEC
        + embodiment  * EMBODIMENT_VEC
        + environment * ENVIRONMENT_VEC
    )


def unit_vector(vec: np.ndarray) -> np.ndarray:
    norm = np.linalg.norm(vec)
    return vec / norm if norm else vec


# =============================================================================
# Ticks - coordinates are semantic; draw locations use axis_pos(...)
# =============================================================================

CONTROLLER_TICKS = [
    (0.00, "No\ncontroller"),
    (0.25, "Oscillators,\nFinite-state\nmachines"),
    (0.50, "Leaky\ninteg.\nneuron"),
    (0.75, "Integ.\nand Fire\nneuron"),
    (1.00, "H.H.\nneuron"),
]

ENVIRONMENT_TICKS = [
    (0.00, "Low"),
    (0.50, "Medium"),
    (1.00, "High"),
]

ACTUATOR_TICKS = [
    (0.00, "No actuator"),
    (0.25, "Servomotor"),
    (0.50, "Torque control"),
    (0.75, "Ekeberg-muscle"),
    (1.00, "Hill-muscle"),
]

NDOF_TICKS = [
    (0.00, "0"),
    (0.20, "10"),
    (0.40, "20"),
    (0.60, "30"),
    (0.80, "40"),
    (1.00, "50+"),
]

# Categorical equivalences. These are used only to snap models to the existing
# original ticks; no new visible ticks are added.
CONTROLLER_LEVEL_TO_COORD = {
    "No controller"               : 0.00,
    "Open-loop / FSM / oscillator": 0.25,
    "Rate / leaky / CPG neural"   : 0.50,
    "ANN / RL / torque policy"    : 0.50,
    "Integrate-and-fire"          : 0.75,
    "Hodgkin-Huxley"              : 1.00,
}

ACTUATOR_LEVEL_TO_COORD = {
    "No actuator"               : 0.00,
    "Servomotor / position"     : 0.25,
    "DC motor / PD"             : 0.25,
    "Torque control"            : 0.50,
    "Ekeberg muscle"            : 0.75,
    "Spring-damper muscle-like" : 0.75,
    "Hill muscle"               : 1.00,
}


# =============================================================================
# Data loading
# =============================================================================

def to_bool(value) -> bool:
    return str(value).strip().lower() in {"yes", "true", "1", "y"}


def to_float(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except ValueError:
        return default


def load_models(path: Path = MODEL_XLSX) -> list[dict]:
    """Load model coordinates from the Models sheet.

    Controller and actuator are discrete variables. Therefore, their plotted
    positions are derived from controller_level and actuator_level and snapped
    to the existing original ticks. The editable coordinate columns remain in
    the workbook for transparency, but they do not create extra ticks.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Models"]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    index = {name: i for i, name in enumerate(headers)}

    models = []
    for raw in rows[1:]:
        row = {name: raw[i] if i < len(raw) else None for name, i in index.items()}

        if not to_bool(row.get("include_in_plot", "yes")):
            continue

        controller_level = str(row.get("controller_level", ""))
        actuator_level   = str(row.get("actuator_level", ""))

        controller_coord = CONTROLLER_LEVEL_TO_COORD.get(
            controller_level,
            to_float(row.get("controller_coord")),
        )
        actuator_coord = ACTUATOR_LEVEL_TO_COORD.get(
            actuator_level,
            to_float(row.get("actuator_coord")),
        )

        color = ROBOT_COLOR if row.get("plot_color") == "red" else MODEL_COLOR
        models.append(
            {
                "label"            : str(row["plot_label"]),
                "controller_level" : controller_level,
                "actuator_level"   : actuator_level,
                "is_robot"         : to_bool(row["is_robot"]),
                "color"            : color,
                "controller"       : controller_coord,
                "actuator"         : actuator_coord,
                "ndof"             : to_float(row["ndof_coord"]),
                "environment"      : to_float(row["environment_coord"]),
                "label_dx"         : to_float(row.get("label_dx"), 0.15),
                "label_dy"         : to_float(row.get("label_dy"), 0.20),
            }
        )

    return models


# =============================================================================
# Drawing utilities
# =============================================================================

def draw_arrow(ax, start: np.ndarray, vec: np.ndarray, color: str) -> None:
    ax.add_patch(
        FancyArrowPatch(
            tuple(start),
            tuple(start + vec),
            arrowstyle     = "-|>",
            mutation_scale = 14,
            color          = color,
            lw             = AXIS_LINEWIDTH,
            shrinkA        = 0,
            shrinkB        = 0,
            zorder         = 10,
        )
    )


def draw_tick(
    ax,
    axis_start: np.ndarray,
    axis_vec: np.ndarray,
    position: float,
    color: str,
    length: float = 0.16,
    lw: float = 1.25,
    zorder: int = 9,
) -> np.ndarray:
    center = axis_start + position * axis_vec
    normal = unit_vector(np.array([-axis_vec[1], axis_vec[0]]))
    a = center - 0.5 * length * normal
    b = center + 0.5 * length * normal
    ax.plot([a[0], b[0]], [a[1], b[1]], color=color, lw=lw, zorder=zorder)
    return center


def box_vertices() -> dict[str, np.ndarray]:
    return {
        "000" : ORIGIN,
        "100" : ORIGIN + EMBODIMENT_VEC,
        "010" : ORIGIN + CONTROLLER_VEC,
        "110" : ORIGIN + EMBODIMENT_VEC + CONTROLLER_VEC,
        "001" : ORIGIN + ENVIRONMENT_VEC,
        "101" : ORIGIN + EMBODIMENT_VEC + ENVIRONMENT_VEC,
        "011" : ORIGIN + CONTROLLER_VEC + ENVIRONMENT_VEC,
        "111" : ORIGIN + EMBODIMENT_VEC + CONTROLLER_VEC + ENVIRONMENT_VEC,
    }


def draw_box_faces(ax) -> None:
    v = box_vertices()
    faces = [
        [v["000"], v["100"], v["101"], v["001"]],
        [v["010"], v["110"], v["111"], v["011"]],
        [v["000"], v["010"], v["011"], v["001"]],
        [v["100"], v["110"], v["111"], v["101"]],
        [v["001"], v["101"], v["111"], v["011"]],
    ]

    for face, alpha in zip(faces, [0.035, 0.035, 0.025, 0.025, 0.045]):
        ax.add_patch(
            Polygon(
                face,
                closed    = True,
                facecolor = "white",
                edgecolor = "none",
                alpha     = alpha,
                zorder    = 5,
            )
        )


def draw_box_edges(ax) -> None:
    v = box_vertices()
    edges = [
        ("000", "100"), ("100", "110"), ("110", "010"), ("010", "000"),
        ("001", "101"), ("101", "111"), ("111", "011"), ("011", "001"),
        ("000", "001"), ("100", "101"), ("010", "011"), ("110", "111"),
    ]

    for a, b in edges:
        p0, p1 = v[a], v[b]
        ax.plot(
            [p0[0], p1[0]],
            [p0[1], p1[1]],
            color  = BOX_COLOR,
            lw     = BOX_LINEWIDTH,
            ls     = (0, (4, 3)),
            alpha  = 0.75,
            zorder = 6,
        )


def draw_sphere(
    ax,
    x: float,
    y: float,
    radius: float = 0.082,
    color: str = MODEL_COLOR,
    alpha: float = SPHERE_ALPHA,
) -> None:
    base_color = np.array(to_rgb(color))
    n_pixels = 90
    yy, xx = np.mgrid[-1:1:complex(n_pixels), -1:1:complex(n_pixels)]
    r2 = xx**2 + yy**2
    mask = r2 <= 1.0

    zz = np.zeros_like(xx)
    zz[mask] = np.sqrt(1.0 - r2[mask])

    light = unit_vector(np.array([-0.55, 0.65, 0.60]))
    normals = np.dstack([xx, yy, zz])
    lambert = np.maximum(0.0, normals @ light)
    highlight = np.exp(-((xx + 0.38) ** 2 + (yy - 0.42) ** 2) / 0.028)

    shade = 0.36 + 0.64 * lambert
    rgb = base_color[None, None, :] * shade[:, :, None]
    rgb = np.clip(rgb + 0.55 * highlight[:, :, None], 0.0, 1.0)

    rim = np.clip((1.0 - r2) / 0.18, 0.0, 1.0)
    rgba = np.zeros((n_pixels, n_pixels, 4))
    rgba[:, :, :3] = rgb
    rgba[:, :, 3] = alpha * mask.astype(float) * (0.55 + 0.45 * rim)

    ax.imshow(
        rgba,
        extent        = [x - radius, x + radius, y - radius, y + radius],
        origin        = "lower",
        interpolation = "bilinear",
        zorder        = 12,
    )


# =============================================================================
# Models and projections
# =============================================================================

def add_overlap_offsets(entries: list[dict]) -> list[dict]:
    """Return entries unchanged.

    Controller and actuator are discrete dimensions. Models sharing the same
    actuator type must therefore sit on the same actuator coordinate. Label
    overlap is handled only by moving text labels, not by jittering model data.
    """
    return [entry.copy() for entry in entries]


def data_delta_from_pixels(ax, dx: float, dy: float) -> np.ndarray:
    """Convert a display-space shift in pixels into data coordinates."""
    inv = ax.transData.inverted()
    p0 = inv.transform((0, 0))
    p1 = inv.transform((dx, dy))
    return p1 - p0


def relax_text_labels(
    fig,
    ax,
    texts: list,
    iterations: int = 220,
    step_px: float = 6.0,
) -> None:
    """Iteratively separate overlapping model labels.

    The model coordinates and projections are not moved.
    """
    if not texts:
        return

    for _ in range(iterations):
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        moved = False

        boxes = [text.get_window_extent(renderer=renderer).expanded(1.06, 1.12)
                 for text in texts]

        for i in range(len(texts)):
            for j in range(i + 1, len(texts)):
                if not boxes[i].overlaps(boxes[j]):
                    continue

                ci = boxes[i].get_points().mean(axis=0)
                cj = boxes[j].get_points().mean(axis=0)
                direction = ci - cj
                norm = np.linalg.norm(direction)

                if norm == 0:
                    direction = np.array([0.0, 1.0])
                else:
                    direction = direction / norm

                direction = direction + np.array([0.18, 0.10])
                norm = np.linalg.norm(direction)
                direction = direction / norm if norm else direction

                delta = data_delta_from_pixels(
                    ax,
                    step_px * direction[0],
                    step_px * direction[1],
                )

                xi, yi = texts[i].get_position()
                xj, yj = texts[j].get_position()

                texts[i].set_position((xi + delta[0], yi + delta[1]))
                texts[j].set_position((xj - delta[0], yj - delta[1]))

                moved = True

        if not moved:
            break


def draw_model(ax, entry: dict):
    """Draw one model and exactly three horizontal projections."""
    # Controller and actuator snap exactly to original discrete tick coordinates.
    controller = axis_pos(entry["controller"])
    actuator   = axis_pos(entry["actuator"])

    # N DOF and environment remain continuous within their axes.
    ndof       = axis_pos(entry["ndof"])
    env        = axis_pos(entry["environment"])

    floor_point = project_point(controller, actuator, 0.0)
    real_point  = project_point(controller, actuator, env)

    controller_axis_point = project_point(controller, 0.0, 0.0)

    ndof_readout_start     = ORIGIN + CONTROLLER_VEC
    actuator_readout_start = ndof_readout_start - NDOF_OFFSET * CONTROLLER_VEC

    # Actuator projection stops exactly on the lower Actuator readout axis
    # at the same discrete actuator tick used by the model itself.
    actuator_readout_point = actuator_readout_start + actuator * EMBODIMENT_VEC

    # N DOF projection points to the original N DOF axis and may fall between ticks.
    ndof_axis_point = NDOF_START + ndof * EMBODIMENT_VEC

    projection_style = {
        "lw"     : PROJECTION_LINEWIDTH,
        "ls"     : (0, (1.2, 2.5)),
        "alpha"  : PROJECTION_ALPHA,
        "zorder" : 2,
    }

    ax.plot(
        [controller_axis_point[0], floor_point[0]],
        [controller_axis_point[1], floor_point[1]],
        color = CONTROLLER_COLOR,
        **projection_style,
    )
    ax.plot(
        [actuator_readout_point[0], floor_point[0]],
        [actuator_readout_point[1], floor_point[1]],
        color = ACTUATOR_COLOR,
        **projection_style,
    )
    ax.plot(
        [ndof_axis_point[0], floor_point[0]],
        [ndof_axis_point[1], floor_point[1]],
        color = NDOF_COLOR,
        **projection_style,
    )

    ax.plot(
        [floor_point[0], real_point[0]],
        [floor_point[1], real_point[1]],
        color  = ENVIRONMENT_COLOR,
        lw     = VERTICAL_LINEWIDTH,
        alpha  = VERTICAL_LINE_ALPHA,
        zorder = 4,
    )

    ax.scatter(
        [floor_point[0]],
        [floor_point[1]],
        marker    = "D",
        s         = 28,
        facecolor = "white",
        edgecolor = entry["color"],
        linewidth = 1.05,
        zorder    = 9,
    )

    draw_sphere(ax, real_point[0], real_point[1], color=entry["color"])

    text = ax.text(
        real_point[0] + entry["label_dx"],
        real_point[1] + entry["label_dy"],
        entry["label"],
        fontsize   = 8.2,
        color      = entry["color"],
        fontweight = "bold" if entry["is_robot"] else "normal",
        ha         = "left",
        va         = "center",
        bbox       = dict(facecolor="white", edgecolor="none", alpha=0.72, pad=0.15),
        zorder     = 13,
    )

    return text


# =============================================================================
# Axes and labels
# =============================================================================

def draw_axes(ax) -> None:
    draw_arrow(ax, ORIGIN,     EMBODIMENT_VEC,  ACTUATOR_COLOR)
    draw_arrow(ax, NDOF_START, EMBODIMENT_VEC,  NDOF_COLOR)
    draw_arrow(ax, ORIGIN,     CONTROLLER_VEC,  CONTROLLER_COLOR)
    draw_arrow(ax, ORIGIN,     ENVIRONMENT_VEC, ENVIRONMENT_COLOR)


def draw_axis_titles(ax) -> None:
    ax.text(
        *(ORIGIN + ENVIRONMENT_VEC + np.array([-0.22, -0.20])),
        "Richness of\nenvironment",
        ha         = "right",
        va         = "bottom",
        fontsize   = 15.0,
        fontweight = "bold",
        color      = ENVIRONMENT_COLOR,
    )

    ax.text(
        *(ORIGIN + CONTROLLER_VEC + np.array([-1.20, -0.60])),
        "Controller",
        ha         = "left",
        va         = "center",
        fontsize   = 15.0,
        fontweight = "bold",
        color      = CONTROLLER_COLOR,
    )

    # Three-line Embodiment title, shifted far right to avoid model labels.
    anchor = ORIGIN + EMBODIMENT_VEC + np.array([1.25, 0.90])

    ax.text(
        *anchor,
        "Embodiment:",
        ha         = "left",
        va         = "center",
        fontsize   = 14.2,
        fontweight = "bold",
        color      = MODEL_COLOR,
        zorder     = 13,
    )
    ax.text(
        *(anchor + np.array([0.0, -0.36])),
        "Actuator",
        ha         = "left",
        va         = "center",
        fontsize   = 14.2,
        fontweight = "bold",
        color      = ACTUATOR_COLOR,
        zorder     = 13,
    )
    ax.text(
        *(anchor + np.array([0.0, -0.72])),
        "N DOF",
        ha         = "left",
        va         = "center",
        fontsize   = 14.2,
        fontweight = "bold",
        color      = NDOF_COLOR,
        zorder     = 13,
    )


def draw_environment_ticks(ax) -> None:
    for coordinate, label in ENVIRONMENT_TICKS:
        tick = draw_tick(ax, ORIGIN, ENVIRONMENT_VEC, axis_pos(coordinate), ENVIRONMENT_COLOR)
        ax.text(
            tick[0] - 0.16,
            tick[1],
            label,
            ha       = "right",
            va       = "center",
            fontsize = 11.0,
            color    = ENVIRONMENT_COLOR,
        )


def draw_controller_ticks(ax) -> None:
    outside_normal = unit_vector(np.array([-CONTROLLER_VEC[1], CONTROLLER_VEC[0]]))
    offsets = [0.12, 0.06, 0.00, -0.05, -0.10]
    for (coordinate, label), dy in zip(CONTROLLER_TICKS, offsets):
        tick = draw_tick(ax, ORIGIN, CONTROLLER_VEC, axis_pos(coordinate), CONTROLLER_COLOR)
        label_position = tick - 0.36 * outside_normal + np.array([-0.05, dy])
        ax.text(
            label_position[0],
            label_position[1],
            label,
            ha       = "right",
            va       = "center",
            fontsize = 8.3,
            color    = CONTROLLER_COLOR,
            zorder   = 13,
        )


def draw_embodiment_axis_ticks(ax) -> None:
    for coordinate, _ in ACTUATOR_TICKS:
        draw_tick(ax, ORIGIN, EMBODIMENT_VEC, axis_pos(coordinate), ACTUATOR_COLOR, length=0.15)
    for coordinate, _ in NDOF_TICKS:
        draw_tick(ax, NDOF_START, EMBODIMENT_VEC, axis_pos(coordinate), NDOF_COLOR, length=0.14)


def draw_readout_axes(ax) -> None:
    ndof_start     = ORIGIN + CONTROLLER_VEC
    actuator_start = ndof_start - NDOF_OFFSET * CONTROLLER_VEC

    for start, color in [(actuator_start, ACTUATOR_COLOR), (ndof_start, NDOF_COLOR)]:
        ax.plot(
            [start[0], start[0] + EMBODIMENT_VEC[0]],
            [start[1], start[1] + EMBODIMENT_VEC[1]],
            color  = color,
            lw     = READOUT_AXIS_LINEWIDTH,
            zorder = 8,
        )

    for coordinate, label in ACTUATOR_TICKS:
        tick = draw_tick(ax, actuator_start, EMBODIMENT_VEC, axis_pos(coordinate), ACTUATOR_COLOR, length=0.15)
        ax.text(
            tick[0],
            tick[1] + 0.18,
            label,
            ha       = "center",
            va       = "bottom",
            fontsize = 8.1,
            color    = ACTUATOR_COLOR,
            zorder   = 13,
        )

    for coordinate, label in NDOF_TICKS:
        tick = draw_tick(ax, ndof_start, EMBODIMENT_VEC, axis_pos(coordinate), NDOF_COLOR, length=0.14)
        ax.text(
            tick[0],
            tick[1] - 0.18,
            label,
            ha       = "center",
            va       = "top",
            fontsize = 8.4,
            color    = NDOF_COLOR,
            zorder   = 13,
        )


def draw_axis_ticks_and_labels(ax) -> None:
    draw_environment_ticks(ax)
    draw_controller_ticks(ax)
    draw_embodiment_axis_ticks(ax)
    draw_readout_axes(ax)


# =============================================================================
# Main
# =============================================================================

def make_figure(models: list[dict] | None = None) -> plt.Figure:
    if models is None:
        models = load_models()

    fig, ax = plt.subplots(figsize=FIGSIZE, dpi=DPI)

    ax.set_aspect("equal", adjustable="box")
    ax.set_xlim(*X_LIMITS)
    ax.set_ylim(*Y_LIMITS)
    ax.axis("off")

    label_texts = []
    for entry in add_overlap_offsets(models):
        label_texts.append(draw_model(ax, entry))

    draw_box_faces(ax)
    draw_box_edges(ax)
    draw_axes(ax)
    draw_axis_titles(ax)
    draw_axis_ticks_and_labels(ax)

    relax_text_labels(fig, ax, label_texts)

    return fig


def main() -> None:
    fig = make_figure()
    fig.savefig(OUT_PNG, bbox_inches="tight", pad_inches=0.08)
    fig.savefig(OUT_PDF, bbox_inches="tight", pad_inches=0.08)
    print(f"Loaded models from: {MODEL_XLSX}")
    print(f"Saved PNG: {OUT_PNG}")
    print(f"Saved PDF: {OUT_PDF}")


if __name__ == "__main__":
    main()
